import streamlit as st
import pandas as pd
from io import BytesIO
from barcode import Code128
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.graphics.barcode import code128 as rl_code128
from reportlab.lib.units import mm
import tempfile
import os

st.set_page_config(page_title="Barcodes erstellen", layout="wide")
st.title("üì¶ Selektive Inventurhilfe")

# üìù Anleitung zur Dateigenerierung aus RWWS
with st.expander("‚ÑπÔ∏è Anleitung: Datei aus dem RWWS exportieren"):
    st.markdown("""
### üìÑ So exportieren Sie Ihre Bestands√ºbersicht aus RWWS

#### üîç 1. RWWS √∂ffnen:
Gehen Sie im **RWWS** zu:  
`Logistik > Warenbewegungen > Operative Listen > Bestands√ºbersicht`

#### üì¶ 2. Selektion einstellen:
- Unter **‚ÄûEinfache Selektion‚Äú**:
  - Warengruppe: **52** (f√ºr Pflanzen)

- Unter **‚ÄûErweiterte Selektion‚Äú**:
  - Beim Feld **‚ÄûGesamtwert (EK)‚Äú** klicken Sie auf die **wei√üe Raute (‚óá)**
  - W√§hlen Sie den **Operator ‚Äû‚â†‚Äú (ungleich)**
  - Geben Sie den Wert **0** ein

#### üì§ 3. Exportieren:
- Klicken Sie auf **‚ÄûDaten Anzeigen‚Äú**
- Danach auf **‚ÄûExport‚Äú**
- W√§hlen Sie **Export nach Excel** und speichern Sie die Datei auf Ihrem PC

#### üì• 4. In die App hochladen:
- Diese Anwendung Aufrufen
- Klicken Sie auf **‚ÄûLaden Sie Ihre Bestands√ºbersicht-Datei hoch (.xlsx)‚Äù**
- W√§hlen Sie Ihre gespeicherte Datei aus
- Die Datei wird automatisch eingelesen und weiterverarbeitet
""")

# üìÅ Datei-Upload
uploaded_file = st.file_uploader("Laden Sie Ihre Bestands√ºbersicht-Datei hoch (.xlsx)", type=["xlsx"])

# üì§ Ausgabeformat w√§hlen
output_option = st.radio("üì§ W√§hlen Sie das Ausgabeformat:", [
    "Excel mit Barcode-Bild",
    "Excel mit Barcode-Text (f√ºr Code128-Schrift)",
    "PDF mit Barcodes (w√§hlbares Format)"
])

# üìê Layoutoption (nur f√ºr PDF)
pdf_layout = None
if output_option == "PDF mit Barcodes (w√§hlbares Format)":
    pdf_layout = st.radio("üìê Seitenlayout f√ºr PDF:", ["Querformat", "Hochformat"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    spalten_zum_entfernen = ["MTART", "Abt.", "WGR", "WGR-Bezeichnung", "Wertart."]
    df = df.drop(columns=[s for s in spalten_zum_entfernen if s in df.columns])

    st.success("‚úÖ Datei erfolgreich geladen.")
    st.dataframe(df.head())

    if st.button("üöÄ Datei erzeugen"):
        def encode_code128(text):
            return chr(204) + text + chr(206)

        if output_option == "Excel mit Barcode-Bild":
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcodes als Bild"
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
            for i, art_nr in enumerate(df["Art-Nr"].astype(str), start=2):
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                    Code128(art_nr, writer=ImageWriter()).write(tmp_img, options={
                        "module_width": 0.25,
                        "module_height": 10,
                        "text_distance": 0,
                        "quiet_zone": 1
                    })
                    tmp_img.close()
                    img = XLImage(tmp_img.name)
                    img.width = 120
                    img.height = 30
                    ws.add_image(img, f"K{i}")
            ws.column_dimensions["K"].width = 22
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="üì• Excel mit Barcode-Bild herunterladen",
                data=output,
                file_name="Bestandsliste_mit_Barcodebildern.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        elif output_option == "Excel mit Barcode-Text (f√ºr Code128-Schrift)":
            df["Barcode"] = df["Art-Nr"].astype(str).apply(encode_code128)
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcodes als Text"
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                label="üì• Excel mit Barcode-Text herunterladen",
                data=output,
                file_name="Bestandsliste_mit_Barcodetext.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        elif output_option == "PDF mit Barcodes (w√§hlbares Format)":
            if pdf_layout == "Querformat":
                page_size = landscape(A4)
                barcode_width = 0.5
                barcode_x_offset = 225 * mm
            else:
                page_size = A4
                barcode_width = 0.35
                barcode_x_offset = 145 * mm

            pdf_buffer = BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=page_size)
            width, height = page_size

            x_margin = 10 * mm
            y_margin = 15 * mm
            x = x_margin
            y = height - y_margin
            line_height = 20 * mm
            max_lines_per_page = int((height - 2 * y_margin) // line_height)
            line_count = 0

            col_pos = {
                "Markt": x + 0 * mm,
                "Art-Nr": x + 20 * mm,
                "Art-Bez": x + 45 * mm,
                "Menge": x + 100 * mm,
                "ME": x + 110 * mm,
                "Wert": x + 120 * mm,
                "VK-Wert": x + 135 * mm,
                "Spanne": x + 155 * mm,
                "EK/VK": x + 175 * mm,
                "GLD": x + 190 * mm,
                "Barcode": barcode_x_offset
            }

            def draw_header():
                c.setFont("Helvetica-Bold", 8)
                for key, xpos in col_pos.items():
                    if key != "Barcode":
                        c.drawString(xpos, y, key)

            draw_header()
            y -= line_height
            line_count += 1

            for _, row in df.iterrows():
                if line_count >= max_lines_per_page:
                    c.showPage()
                    y = height - y_margin
                    draw_header()
                    y -= line_height
                    line_count = 1

                c.setFont("Helvetica", 7)
                c.drawString(col_pos["Markt"], y, str(row["Markt"]))
                c.drawString(col_pos["Art-Nr"], y, str(row["Art-Nr"]))
                c.drawString(col_pos["Art-Bez"], y, str(row["Art-Bez"])[:50])
                c.drawRightString(col_pos["Menge"] + 8, y, str(row["Menge"]))
                c.drawString(col_pos["ME"], y, str(row["ME"]))
                c.drawRightString(col_pos["Wert"] + 10, y, f'{row["Wert"]:.2f}')
                c.drawRightString(col_pos["VK-Wert"] + 10, y, f'{row["VK-Wert"]:.2f}')
                c.drawRightString(col_pos["Spanne"] + 10, y, f'{row["Spanne"]:.2f}')
                c.drawRightString(col_pos["EK/VK"] + 10, y, f'{row["EK/VK"]:.3f}')
                c.drawRightString(col_pos["GLD"] + 10, y, f'{row["GLD"]:.2f}')

                barcode = rl_code128.Code128(str(row["Art-Nr"]), barHeight=11 * mm, barWidth=barcode_width)
                barcode.drawOn(c, col_pos["Barcode"], y - 2)

                y -= line_height
                line_count += 1

            c.save()
            pdf_buffer.seek(0)

            st.download_button(
                label=f"üì• PDF ({pdf_layout}) herunterladen",
                data=pdf_buffer,
                file_name=f"Bestandsliste_Barcodes_{pdf_layout}.pdf",
                mime="application/pdf"
            )

# Footer Hinweis
st.markdown("""
---
‚ö†Ô∏è Hinweis: Diese Anwendung speichert keine Daten und hat keinen Zugriff auf Ihre Dateien.

üåü *Erstellt von Christoph R. Kaiser mit Hilfe von K√ºnstlicher Intelligenz.*
""")
